import io
import os
import re
import time
import shutil
import pandas as pd
import ipywidgets as widgets
from IPython.display import display, clear_output
from openpyxl import load_workbook
from google.colab import drive

# --- 0. GOOGLE DRIVE CONNECTION ---
print("🔗 Connecting to Google Drive...")
try:
    # Increased timeout slightly to handle busy multi-user channels safely
    drive.mount('/content/drive', timeout_ms=30000)
    DRIVE_AVAILABLE = True
except Exception as e:
    print("⚠️ Google Drive mount skipped. Using local runtime storage.")
    DRIVE_AVAILABLE = False

DRIVE_PROJECT_FOLDER = '/content/drive/MyDrive/Python Project/' if DRIVE_AVAILABLE else './'

MASTER_CATALOG_FILENAME = "Catalog Master.xlsx"
MASTER_CATALOG_PATH = os.path.join(DRIVE_PROJECT_FOLDER, MASTER_CATALOG_FILENAME)
BUYER_FILE_NAME = "Buyer File.xlsx"
BUYER_FILE_PATH = os.path.join(DRIVE_PROJECT_FOLDER, BUYER_FILE_NAME)
LOCK_FILE_PATH = os.path.join(DRIVE_PROJECT_FOLDER, "drive_write.lock")

if not os.path.exists(MASTER_CATALOG_PATH) or not os.path.exists(BUYER_FILE_PATH):
    raise FileNotFoundError("❌ Missing baseline master dependencies in Drive folder.")

# --- 1. LOAD MASTER PARTNER STRUCTURES ---
wb_master = load_workbook(MASTER_CATALOG_PATH)
sheet_names = wb_master.sheetnames
vendor_sheet_name = "PivVendorMaster" if "PivVendorMaster" in sheet_names else sheet_names[0]
sheet_v = wb_master[vendor_sheet_name]

partner_db = {}
for row in range(1, sheet_v.max_row + 1):
    p_name = sheet_v.cell(row=row, column=1).value
    p_code = sheet_v.cell(row=row, column=3).value
    curr = sheet_v.cell(row=row, column=4).value
    ship_t = sheet_v.cell(row=row, column=6).value
    b_code = sheet_v.cell(row=row, column=16).value
    if p_name is not None:
        p_name_str = str(p_name).strip()
        p_name_upper = p_name_str.upper()
        if p_name_str == "" or "VENDOR_NAME" in p_name_upper or "PLEASE SELECT" in p_name_upper:
            continue
        partner_db[p_name_upper] = {
            "original_name": p_name_str,
            "partner_code": str(p_code).strip() if p_code else "",
            "buyer_code": str(b_code).strip().upper() if b_code else "ZN",
            "currency": str(curr).strip().upper() if curr else "MYR",
            "shipping_term": str(ship_t).strip().upper() if ship_t else "PLEASE SELECT"
        }
wb_master.close()

sorted_partners = sorted(list(set([v["original_name"] for v in partner_db.values()])))
umsr_options = ["BT", "BUL", "BX", "CT", "DAY", "DRM", "GRM", "HR", "JOB", "KGM", "LOT", "LTR", "M3", "MAN", "MLT", "MM", "MTH", "NRL", "PCE", "PKT", "PR", "PT", "SET", "SHT", "TON", "TRP", "UT", "PLEASE SELECT"]
shipping_options = ["CIF", "CIP", "DAP", "DAT", "DDP", "DDU", "EXW", "FCA", "FOB", "OTHERS", "PLEASE SELECT"]

row_widgets_collection = []

# --- Helper: Sequential Sequence Finder ---
def get_highest_sequence_number_from_buyer_file(sheet, prefix):
    highest_num = 0
    r = 3
    empty_row_allowance = 0
    while True:
        cell_val = sheet.cell(row=r, column=2).value
        if cell_val is None or str(cell_val).strip() == "" or "PLEASE SELECT" in str(cell_val).upper():
            empty_row_allowance += 1
            if empty_row_allowance > 15: break
        else:
            empty_row_allowance = 0
            cell_str = str(cell_val).strip().upper()
            # Updated regex to flexibly handle variable whitespace between your prefixes and sequence digits
            match = re.match(rf"^{re.escape(prefix)}\s*(\d+)", cell_str)
            if match:
                val_num = int(match.group(1))
                if val_num > highest_num: highest_num = val_num
        r += 1
    return highest_num

def get_base_fallback(full_prefix):
    fallbacks = {"CP": 239, "CR": 8974, "NP": 12514, "NR": 16802, "CCP": 364, "CCR": 1575, "XP": 58, "XR": 110}
    return fallbacks.get(full_prefix, 2600)

def format_price_two_decimals(val_str):
    try:
        val_clean = str(val_str).strip()
        if not val_clean: return ""
        return f"{float(val_clean):.2f}"
    except:
        return str(val_str).strip()

# --- 2. BASE INTERACTIVE INTERFACE ---
clear_output(wait=True)
print("⚙️ INTERACTIVE DYNAMIC MULTI-ROW CATALOG HUB (MANUAL CORRECTION MODE)")
print("-" * 110)

file_uploader = widgets.FileUpload(accept='.xlsx, .xls, .xlsm', multiple=False, description="Upload Request File", button_style="info", icon="upload")
operator_dropdown = widgets.Dropdown(options=["PLEASE SELECT", "Luqman", "Alisya", "Tan"], value="PLEASE SELECT", description="Buyer:")
code_category_dropdown = widgets.Dropdown(options=["PLEASE SELECT", "Controllable ", "Comparison ", "Chemical", "Non-Controllable "], value="PLEASE SELECT", description="Cat Type:", layout={'width': '400px'})
submit_btn = widgets.Button(description="Confirm & Sync All Rows", button_style="success", icon="check", layout={'width': '320px'})

grid_container = widgets.VBox()
out = widgets.Output()

# Re-evaluate live catalog preview codes if Buyer or Cat Type shifts dropdown options
def update_live_catalog_previews(change):
    global row_widgets_collection
    if operator_dropdown.value == "PLEASE SELECT" or code_category_dropdown.value == "PLEASE SELECT":
        for r_obj in row_widgets_collection:
            r_obj["cat_code_widget"].value = "[SELECT OPTIONS ABOVE]"
        return

    op = operator_dropdown.value
    cat = code_category_dropdown.value
    suffix_char = "R" if op == "Luqman" else ("P" if op == "Alisya" else "D")
    prefix_letter = "C" if cat == "Controllable " else ("CC" if cat == "Comparison " else ("X" if cat == "Chemical" else "N"))
    full_prefix = f"{prefix_letter}{suffix_char}"

    sheet_mapping = {"Controllable ": "Controllable", "Comparison ": "Comparison", "Chemical": "Chemical (X)", "Non-Controllable ": "Non Controllable"}
    target_sheet_name = sheet_mapping.get(cat, "Non Controllable")

    try:
        wb_b_check = load_workbook(BUYER_FILE_PATH, data_only=True)
        highest_seq = 0
        if target_sheet_name in wb_b_check.sheetnames:
            highest_seq = get_highest_sequence_number_from_buyer_file(wb_b_check[target_sheet_name], full_prefix)
        wb_b_check.close()
        if highest_seq == 0: highest_seq = get_base_fallback(full_prefix)

        for idx, r_obj in enumerate(row_widgets_collection):
            highest_seq += 1
            r_obj["cat_code_widget"].value = f"{full_prefix}{highest_seq}"
    except Exception as e:
        pass

operator_dropdown.observe(update_live_catalog_previews, names='value')
code_category_dropdown.observe(update_live_catalog_previews, names='value')

# --- 3. DYNAMIC INTERACTIVE MATRIX INJECTOR ---
def on_file_upload(change):
    global row_widgets_collection
    if not file_uploader.value: return

    try:
        # FIX: Safe extraction mapping compatible with both ipywidgets v7 and v8+ tuple parsing
        if isinstance(file_uploader.value, dict):
            uploaded_file = list(file_uploader.value.values())[0]
        else:
            uploaded_file = file_uploader.value[0]
            
        content = uploaded_file['content']
        wb_upload = load_workbook(io.BytesIO(content), data_only=True)
        target_sheets = [s for s in wb_upload.sheetnames if "PivVendorMaster" not in s]
        sheet = wb_upload[target_sheets[0]]

        col_map = {}
        for col in range(1, sheet.max_column + 1):
            header_val = str(sheet.cell(row=13, column=col).value or "").upper()
            if "PARTNER NAME" in header_val: col_map["partner"] = col
            elif "CATEGORY" in header_val: col_map["category"] = col
            elif "ITEM DESCRIPTION" in header_val and "46" in header_val: col_map["desc"] = col
            elif "PURCHASE UMSR" in header_val: col_map["umsr"] = col
            elif "VALIDITY" in header_val: col_map["validity"] = col
            elif "FINAL U/PRICE" in header_val: col_map["final_price"] = col
            elif "CURRENCY" in header_val: col_map["currency"] = col
            elif "INITIAL" in header_val: col_map["init_price"] = col
            elif "MOQ" in header_val: col_map["moq"] = col
            elif "LEAD TIME" in header_val: col_map["lead"] = col
            elif "SHIPPING TERM" in header_val: col_map["ship_term"] = col

        row_widgets_collection = []
        grid_children = [
            widgets.HTML("<b style='color:#2c3e50; font-size:13px;'>📝 CORRECT EXTRACTED DATA IN-LINE BELOW BEFORE SUBMITTING:</b>"),
            widgets.HTML("<hr style='margin:5px 0;'>")
        ]

        curr_row = 14
        while True:
            test_val = sheet.cell(row=curr_row, column=col_map.get("desc", 3)).value
            if test_val is None or str(test_val).strip() == "":
                break

            def get_row_val(key, default=""):
                if key in col_map:
                    v = sheet.cell(row=curr_row, column=col_map[key]).value
                    return default if v is None else str(v).strip()
                return default

            ext_partner = get_row_val("partner").upper()
            p_details = partner_db.get(ext_partner, {"partner_code": "", "buyer_code": "ZN", "currency": "MYR", "shipping_term": "PLEASE SELECT"})

            raw_price = get_row_val("final_price")
            formatted_price = format_price_two_decimals(raw_price) if raw_price else ""

            raw_init = get_row_val("init_price")
            formatted_init = format_price_two_decimals(raw_init) if raw_init else ""

            w_cat_code = widgets.Text(value="[CHOOSE PROFILE]", disabled=True, layout={'width': '110px'}, style={'description_width': 'initial'})
            w_partner = widgets.Dropdown(options=sorted_partners, value=ext_partner if ext_partner in sorted_partners else sorted_partners[0], layout={'width': '180px'})
            w_cat = widgets.Dropdown(options=["Goods/Item", "Others", "blank"], value=get_row_val("category", "Others"), layout={'width': '100px'})
            w_desc = widgets.Text(value=get_row_val("desc").upper()[:46], layout={'width': '220px'}, placeholder="Description")
            w_pcode = widgets.Text(value=p_details["partner_code"], layout={'width': '90px'}, placeholder="P. Code")
            w_bcode = widgets.Text(value=p_details["buyer_code"], layout={'width': '60px'}, placeholder="Buyer")
            w_umsr = widgets.Dropdown(options=umsr_options, value=get_row_val("umsr", "PCE").upper() if get_row_val("umsr", "PCE").upper() in umsr_options else "PCE", layout={'width': '90px'})
            w_validity = widgets.Text(value=get_row_val("validity").split('.')[0] if "none" not in get_row_val("validity").lower() else "", layout={'width': '90px'}, placeholder="YYYYMMDD")
            w_price = widgets.Text(value=formatted_price, layout={'width': '80px'}, placeholder="Price")
            w_curr = widgets.Text(value=get_row_val("currency", p_details["currency"]).upper(), layout={'width': '60px'})
            w_init = widgets.Text(value=formatted_init, layout={'width': '80px'}, placeholder="Initial")
            w_moq = widgets.Text(value="", layout={'width': '50px'}, placeholder="MOQ")
            w_lead = widgets.Text(value="", layout={'width': '50px'}, placeholder="Lead")
            w_ship = widgets.Dropdown(options=shipping_options, value=get_row_val("ship_term", p_details["shipping_term"]) if get_row_val("ship_term", p_details["shipping_term"]) in shipping_options else "PLEASE SELECT", layout={'width': '120px'})

            def on_price_blur(change, text_widget=w_price):
                text_widget.value = format_price_two_decimals(text_widget.value)
            w_price.observe(on_price_blur, names='value')

            def on_init_blur(change, text_widget=w_init):
                text_widget.value = format_price_two_decimals(text_widget.value)
            w_init.observe(on_init_blur, names='value')

            def link_partner_change(change, pc_w=w_pcode, bc_w=w_bcode, cur_w=w_curr, sh_w=w_ship):
                p_selected = change['new'].upper()
                if p_selected in partner_db:
                    pc_w.value = partner_db[p_selected]["partner_code"]
                    bc_w.value = partner_db[p_selected]["buyer_code"]
                    cur_w.value = partner_db[p_selected]["currency"]
                    sh_w.value = partner_db[p_selected]["shipping_term"] if partner_db[p_selected]["shipping_term"] in shipping_options else "PLEASE SELECT"

            w_partner.observe(link_partner_change, names='value')

            row_label = widgets.HTML(f"<b style='color:#e74c3c;'>Row {curr_row}:</b> ", layout={'width': '50px'})
            hbox_row = widgets.HBox([
                row_label, w_partner, w_cat, w_desc, w_cat_code, w_pcode, w_bcode, w_umsr, w_validity, w_price, w_curr, w_init, w_moq, w_lead, w_ship
            ])

            grid_children.append(hbox_row)
            row_widgets_collection.append({
                "partner": w_partner, "category": w_cat, "desc": w_desc, "cat_code_widget": w_cat_code, "pcode": w_pcode, "bcode": w_bcode,
                "umsr": w_umsr, "validity": w_validity, "price": w_price, "currency": w_curr, "init_price": w_init,
                "moq": w_moq, "lead": w_lead, "ship_term": w_ship
            })
            curr_row += 1

        wb_upload.close()
        grid_container.children = grid_children
        update_live_catalog_previews(None)
        print(f"✨ Parsed {len(row_widgets_collection)} items into editable matrix forms below.")

    except Exception as e:
        print(f"❌ Core Scanning Failure: {e}")

file_uploader.observe(on_file_upload, names='value')

# --- 4. DATA SYNCHRONIZATION AND SAVE ENGINE (CONCURRENCY-PROTECTED) ---
def on_click_sync_all(b):
    global row_widgets_collection
    with out:
        clear_output()
        if not row_widgets_collection:
            print("❌ Processing Error: No rows staged. Please attach your request spreadsheet first.")
            return
        if operator_dropdown.value == "PLEASE SELECT" or code_category_dropdown.value == "PLEASE SELECT":
            print("❌ Selection Error: Assigned Buyer and Cat Type options must be explicitly specified.")
            return

        # === 🔒 LIVE QUEUE CONCURRENCY LOCK LOCKOUT ===
        retry_count = 0
        while os.path.exists(LOCK_FILE_PATH):
            print("⏳ Another buyer is sync saving data right now. Auto-waiting for queue clear...")
            time.sleep(2)
            retry_count += 1
            if retry_count > 10:
                print("⚠️ Sync Error: Google Drive channel is heavily congested. Please click button again.")
                return

        with open(LOCK_FILE_PATH, 'w') as lock_f:
            lock_f.write("locked")
        # ===============================================

        try:
            op = operator_dropdown.value
            cat = code_category_dropdown.value
            suffix_char = "R" if op == "Luqman" else ("P" if op == "Alisya" else "D")
            prefix_letter = "C" if cat == "Controllable " else ("CC" if cat == "Comparison " else ("X" if cat == "Chemical" else "N"))
            full_prefix = f"{prefix_letter}{suffix_char}"

            sheet_mapping = {"Controllable ": "Controllable", "Comparison ": "Comparison", "Chemical": "Chemical (X)", "Non-Controllable ": "Non Controllable"}
            target_sheet_name = sheet_mapping.get(cat, "Non Controllable")

            print("💾 Safely opening master data from Google Drive repository...")
            
            # FRESH RE-READ AFTER ACQUIRING LOCK TO GRAB THE ABSOLUTE LATEST SEQUENCE STAMPED BY OTHER USERS
            wb_b_check = load_workbook(BUYER_FILE_PATH, data_only=True)
            highest_seq = 0
            if target_sheet_name in wb_b_check.sheetnames:
                highest_seq = get_highest_sequence_number_from_buyer_file(wb_b_check[target_sheet_name], full_prefix)
            wb_b_check.close()
            if highest_seq == 0: highest_seq = get_base_fallback(full_prefix)

            wb_m = load_workbook(MASTER_CATALOG_PATH)
            m_sheets = [s for s in wb_m.sheetnames if s != "PivVendorMaster"]
            sheet_m = wb_m[m_sheets[0]]

            # Explicit row discovery targeting the actual absolute bottom of sheets
            m_row = sheet_m.max_row + 1
            wb_b = load_workbook(BUYER_FILE_PATH)
            sheet_b = wb_b[target_sheet_name] if target_sheet_name in wb_b.sheetnames else wb_b.create_sheet(title=target_sheet_name)
            b_row = sheet_b.max_row + 1

            def clean_int(val): return int(val) if str(val).strip().isdigit() else None
            def clean_float(val):
                try: return float(val)
                except: return None
            def safe_w(sheet, r, c, val):
                if type(sheet.cell(row=r, column=c)).__name__ != 'MergedCell':
                    sheet.cell(row=r, column=c).value = val

            print(f"🚀 Syncing manual overrides sequentially starting at Catalog Row {m_row}, Buyer Row {b_row}:")
            for idx, r_obj in enumerate(row_widgets_collection):
                highest_seq += 1
                generated_code = f"{full_prefix}{highest_seq}"

                corrected_partner = r_obj["partner"].value.upper()
                corrected_desc = r_obj["desc"].value.strip().upper()[:46]

                # Map directly to Catalog Master
                safe_w(sheet_m, m_row, 1, max(1, m_row - 13))
                safe_w(sheet_m, m_row, 2, corrected_partner)
                safe_w(sheet_m, m_row, 3, r_obj["category"].value)
                safe_w(sheet_m, m_row, 4, corrected_desc)
                safe_w(sheet_m, m_row, 5, generated_code)
                safe_w(sheet_m, m_row, 6, "RST")
                safe_w(sheet_m, m_row, 7, corrected_desc)
                safe_w(sheet_m, m_row, 8, r_obj["pcode"].value.upper())
                safe_w(sheet_m, m_row, 9, r_obj["bcode"].value.upper())
                safe_w(sheet_m, m_row, 10, r_obj["umsr"].value)
                safe_w(sheet_m, m_row, 11, 0)
                safe_w(sheet_m, m_row, 12, "")
                safe_w(sheet_m, m_row, 13, r_obj["validity"].value)
                safe_w(sheet_m, m_row, 14, clean_float(r_obj["price"].value))
                safe_w(sheet_m, m_row, 15, r_obj["currency"].value.upper())
                safe_w(sheet_m, m_row, 16, clean_float(r_obj["init_price"].value))
                safe_w(sheet_m, m_row, 17, clean_int(r_obj["moq"].value))
                safe_w(sheet_m, m_row, 18, "")
                safe_w(sheet_m, m_row, 19, clean_int(r_obj["lead"].value))
                safe_w(sheet_m, m_row, 20, r_obj["ship_term"].value)

                # Map to Buyer File
                safe_w(sheet_b, b_row, 1, "")
                safe_w(sheet_b, b_row, 2, generated_code)
                safe_w(sheet_b, b_row, 3, corrected_desc)
                safe_w(sheet_b, b_row, 4, corrected_partner)

                print(f"  🔹 Row added successfully: {generated_code} -> {corrected_desc}")
                m_row += 1
                b_row += 1

            # STAGING BUFFERS: Write locally first before atomic transfer swap over network
            temp_catalog_path = MASTER_CATALOG_PATH + ".tmp"
            temp_buyer_path = BUYER_FILE_PATH + ".tmp"
            
            wb_m.save(temp_catalog_path)
            wb_b.save(temp_buyer_path)
            
            wb_m.close()
            wb_b.close()
            
            shutil.move(temp_catalog_path, MASTER_CATALOG_PATH)
            shutil.move(temp_buyer_path, BUYER_FILE_PATH)

            print(f"\n🎉 Success! All row data corrected in-line has been safely committed to Google Drive.")
            row_widgets_collection = []
            grid_container.children = []

        except Exception as tx_ex:
            print(f"❌ Database Transaction Error: {tx_ex}")
            
        finally:
            if os.path.exists(LOCK_FILE_PATH):
                os.remove(LOCK_FILE_PATH)

submit_btn.on_click(on_click_sync_all)

# --- 5. RENDER SYSTEM MATRIX LAYOUT ---
# FIX: Removed invalid trailing ellipsis statement causing compilation crashes
display(
    widgets.HTML("<h3>📂 1. UPLOAD USER REQUEST FILE</h3>"), file_uploader,
    widgets.HTML("<h3>👤 2. ASSIGN CORE BUYER PROFILE & CLASS</h3>"), operator_dropdown, code_category_dropdown,
    widgets.HTML("<hr>"), grid_container, widgets.HTML("<br>"), submit_btn, out
)
   
